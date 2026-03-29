# extractor.py
# Ce fichier s'occupe d'extraire le texte brut depuis différents formats de fichiers.
# On utilise des librairies spécialisées selon le type du fichier.

import os

def extraire_texte(chemin_fichier):
    """
    Reçoit le chemin d'un fichier et retourne son contenu textuel.
    Supporte : PDF, Word (.docx), Excel (.xlsx), et fichiers texte (.txt).
    Retourne None si le format n'est pas supporté.
    """
    extension = os.path.splitext(chemin_fichier)[1].lower()

    if extension == ".pdf":
        return extraire_pdf(chemin_fichier)
    elif extension == ".docx":
        return extraire_word(chemin_fichier)
    elif extension == ".xlsx":
        return extraire_excel(chemin_fichier)
    elif extension == ".txt":
        return extraire_txt(chemin_fichier)
    else:
        return None  # Format non reconnu


def extraire_pdf(chemin):
    """Extrait le texte d'un fichier PDF page par page."""
    try:
        import fitz  # PyMuPDF
        texte_total = ""
        document = fitz.open(chemin)
        for page in document:
            texte_total += page.get_text()
        document.close()
        return texte_total.strip()
    except ImportError:
        return "[Erreur] La librairie PyMuPDF n'est pas installée. Lancez : pip install pymupdf"
    except Exception as e:
        return f"[Erreur lors de la lecture du PDF] {e}"


def extraire_word(chemin):
    """Extrait le texte d'un fichier Word (.docx) paragraphe par paragraphe."""
    try:
        from docx import Document
        doc = Document(chemin)
        lignes = [paragraphe.text for paragraphe in doc.paragraphs if paragraphe.text.strip()]
        return "\n".join(lignes)
    except ImportError:
        return "[Erreur] La librairie python-docx n'est pas installée. Lancez : pip install python-docx"
    except Exception as e:
        return f"[Erreur lors de la lecture du fichier Word] {e}"


def extraire_excel(chemin):
    """Extrait le contenu des cellules d'un fichier Excel (.xlsx)."""
    try:
        import openpyxl
        classeur = openpyxl.load_workbook(chemin, data_only=True)
        contenu = []
        for nom_feuille in classeur.sheetnames:
            feuille = classeur[nom_feuille]
            contenu.append(f"[Feuille : {nom_feuille}]")
            for ligne in feuille.iter_rows(values_only=True):
                valeurs = [str(cellule) for cellule in ligne if cellule is not None]
                if valeurs:
                    contenu.append(" | ".join(valeurs))
        return "\n".join(contenu)
    except ImportError:
        return "[Erreur] La librairie openpyxl n'est pas installée. Lancez : pip install openpyxl"
    except Exception as e:
        return f"[Erreur lors de la lecture du fichier Excel] {e}"


def extraire_txt(chemin):
    """Lit un fichier texte brut."""
    try:
        with open(chemin, "r", encoding="utf-8") as f:
            return f.read().strip()
    except UnicodeDecodeError:
        # Certains fichiers texte ne sont pas en UTF-8, on essaie latin-1
        try:
            with open(chemin, "r", encoding="latin-1") as f:
                return f.read().strip()
        except Exception as e:
            return f"[Erreur d'encodage] {e}"
    except Exception as e:
        return f"[Erreur lors de la lecture du fichier texte] {e}"
